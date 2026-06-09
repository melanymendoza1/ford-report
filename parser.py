#!/usr/bin/env python3
"""
ORGU FORD — Parser de datos AEADE v2 (Jun 2026)
Uso: python parser.py <ruta_al_excel.xlsx>
Genera: public/report_data.json
"""

import json, sys, os
from datetime import datetime
from openpyxl import load_workbook

PROVINCES = ['PICHINCHA', 'GUAYAS', 'MANABÍ', 'EL ORO']

def safe_num(v):
    if v is None: return None
    try:
        f = float(v)
        return None if f != f else f
    except:
        return None

def is_year(v):
    return v in (2024, 2025, 2026, '2024', '2025', '2026')

# ─── Generic flat-table extractor ─────────────────────────────────────────────
# Reads any sheet that has: filter rows, then "Etiquetas de fila" header row,
# then data rows like (label, 2024, 2025, 2026, Total)
def extract_flat(ws, max_rows=500):
    """Returns list of {label, y2024, y2025, y2026} — skips Total general."""
    rows = list(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True))
    for i, row in enumerate(rows):
        if row[0] == 'Etiquetas de fila' and any(is_year(v) for v in row):
            header = [str(v) for v in row]
            result = []
            for r in rows[i+1:]:
                if r[0] is None: break
                if str(r[0]) == 'Total general': continue
                d = {'label': r[0]}
                for j, h in enumerate(header[1:], 1):
                    if h in ('2024','2025','2026'):
                        d[f'y{h}'] = safe_num(r[j]) if j < len(r) else None
                result.append(d)
            return result
    return []

# ─── Province-segmented extractor ─────────────────────────────────────────────
# Handles sheets where data appears as one flat block after "Etiquetas de fila",
# with province names as data rows (PICHINCHA row → then MAZDA, KIA... then GUAYAS row → ...)
# brand_level=True  → only brand-level rows per province (skip trim rows)
# brand_level=False → all rows labeled, no filtering
TRIM_KEYWORDS = ('AC ','TM ',' TA ','DIESEL','HYBRID','CVT',' HEV','ECOBOOST','TURBO',
                 'CRDI','VTEC','DOHC','MHEV','PHEV','BEV')

def is_trim_row(label):
    if len(label) > 40:
        return True
    # Check keywords, also check if label ends with common transmission codes
    if any(kw in label for kw in TRIM_KEYWORDS):
        return True
    # Trim rows often end with TA, TM, or contain version numbers like "1.5", "2.0"
    if label.endswith(' TA') or label.endswith(' TM'):
        return True
    import re
    if re.search(r'\d+\.\d+ \d+P', label):  # e.g. "1.5 5P" engine/body pattern
        return True
    return False

def extract_by_province(ws, max_rows=800, brand_level=True):
    rows = list(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True))
    result = {}

    # Find the header row with years
    header_idx = None
    year_cols = {}
    for i, row in enumerate(rows):
        if row[0] == 'Etiquetas de fila' and any(is_year(v) for v in row):
            header_idx = i
            for j, h in enumerate(row):
                if is_year(h):
                    year_cols[str(h)] = j
            break

    if header_idx is None:
        return result

    # Walk data rows — province rows have numbers in year columns AND match PROVINCES
    current_prov = None
    for row in rows[header_idx+1:]:
        if row[0] is None:
            continue
        label = str(row[0])
        if label == 'Total general':
            continue

        if label in PROVINCES:
            current_prov = label
            if current_prov not in result:
                result[current_prov] = []
            continue

        if current_prov is None:
            continue

        if brand_level and is_trim_row(label):
            continue  # skip trim rows in brand_level mode

        # Skip dot-separated sub-brand labels like "RAM . RAM 1500" or "FORD . FORD F-150"
        if ' . ' in label:
            continue

        d = {'brand': label} if brand_level else {'label': label}
        for yr, col in year_cols.items():
            d[f'y{yr}'] = safe_num(row[col]) if col < len(row) else None
        result[current_prov].append(d)

    return result

# ─── Nacional brand extractor (no province header) ────────────────────────────
def extract_nacional_brands(ws, max_rows=600):
    """For sheets like SUV_GAS_25_40_INDUSTRIA_YTD — brand → trim hierarchy, no province."""
    rows = list(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True))
    for i, row in enumerate(rows):
        if row[0] == 'Etiquetas de fila' and any(is_year(v) for v in row):
            header = list(row)
            year_cols = {}
            for j, h in enumerate(header):
                if is_year(h):
                    year_cols[str(h)] = j
            brands = []
            j2 = i + 1
            while j2 < len(rows):
                r = rows[j2]
                if r[0] is None: break
                if str(r[0]) == 'Total general':
                    j2 += 1; continue
                label = str(r[0])
                # Skip trim rows
                if is_trim_row(label):
                    j2 += 1; continue
                d = {'brand': label}
                for yr, col in year_cols.items():
                    d[f'y{yr}'] = safe_num(r[col]) if col < len(r) else None
                brands.append(d)
                j2 += 1
            return brands
    return []

# ─── FY/YTD summary extractor ─────────────────────────────────────────────────
def extract_summary(ws):
    """Returns {cat: {y2024, y2025, y2026, fcts2026}} — for FY sheets with forecast."""
    rows = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))
    for i, row in enumerate(rows):
        if row[0] == 'Etiquetas de fila':
            header = list(row)
            result = {}
            for r in rows[i+1:]:
                if r[0] is None: break
                if str(r[0]) == 'Total general':
                    result['_total'] = {
                        'y2024': safe_num(r[1]),
                        'y2025': safe_num(r[2]),
                        'y2026': safe_num(r[3]),
                    }
                    # fcts = (ytd/5)*12
                    if safe_num(r[3]):
                        result['_total']['fcts2026'] = round((safe_num(r[3]) / 5) * 12)
                    continue
                cat = str(r[0])
                result[cat] = {
                    'y2024': safe_num(r[1]),
                    'y2025': safe_num(r[2]),
                    'y2026': safe_num(r[3]),
                }
                if safe_num(r[3]):
                    result[cat]['fcts2026'] = round((safe_num(r[3]) / 5) * 12)
            return result
    return {}

# ─── Province totals extractor ─────────────────────────────────────────────────
def extract_province_totals(ws):
    """For PROVINCIAS_YTD, FORD_PROVINCIAS_YTD — returns {prov: {y2024, y2025, y2026}}."""
    rows = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))
    for i, row in enumerate(rows):
        if row[0] == 'Etiquetas de fila':
            result = {}
            for r in rows[i+1:]:
                if r[0] is None: break
                if str(r[0]) == 'Total general':
                    result['ZONA ORGU'] = {'y2024': safe_num(r[1]), 'y2025': safe_num(r[2]), 'y2026': safe_num(r[3])}
                    continue
                result[str(r[0])] = {'y2024': safe_num(r[1]), 'y2025': safe_num(r[2]), 'y2026': safe_num(r[3])}
            return result
    return {}

# ─── Ford category by province ─────────────────────────────────────────────────
def extract_ford_cat_by_prov(ws):
    """FORD_CATEGORIA_POR_PROV_YTD — {prov: {cat: {y2024, y2025, y2026}}}"""
    rows = list(ws.iter_rows(min_row=1, max_row=100, values_only=True))
    result = {}
    current_prov = None
    for i, row in enumerate(rows):
        if row[0] in PROVINCES:
            current_prov = row[0]
            result[current_prov] = {}
        elif current_prov and row[0] and row[0] not in ('Etiquetas de fila', 'Total general', None):
            label = str(row[0])
            if label not in PROVINCES and not any(is_year(label) for _ in [1]):
                result[current_prov][label] = {
                    'y2024': safe_num(row[1]),
                    'y2025': safe_num(row[2]),
                    'y2026': safe_num(row[3]),
                }
    return result

# ─── Pick Up diesel combined TM+TA ─────────────────────────────────────────────
def extract_pick_diesel_combined(ws_tm, ws_ta):
    """Merges TM and TA provincial brand data."""
    tm = extract_by_province(ws_tm, brand_level=True)
    ta = extract_by_province(ws_ta, brand_level=True)
    combined = {}
    all_provs = set(list(tm.keys()) + list(ta.keys()))
    for prov in all_provs:
        brand_map = {}
        for source in [tm.get(prov, []), ta.get(prov, [])]:
            for item in source:
                b = item['brand']
                if b not in brand_map:
                    brand_map[b] = {'brand': b, 'y2024': 0, 'y2025': 0, 'y2026': 0}
                for yr in ('y2024', 'y2025', 'y2026'):
                    brand_map[b][yr] = (brand_map[b][yr] or 0) + (item.get(yr) or 0)
        combined[prov] = list(brand_map.values())
    return combined

# ─── MAIN ──────────────────────────────────────────────────────────────────────
def reshape_for_frontend(report: dict, wb) -> dict:
    """
    page.tsx espera claves específicas con formato array o nested.
    Aquí hacemos el mapping sin tocar el frontend.
    """
    out = {k: v for k, v in report.items()}  # copy all existing keys

    months = report.get('months_ytd', 5)

    # ── T1: mercado_ytd, ford_ytd (arrays con cat/y2024/y2025/y2026) ──
    # Zona Orgu data (lo que usa el tab principal)
    ind_orgu = report.get('industria_orgu_ytd', {})
    ford_orgu = report.get('ford_orgu_ytd', {})
    ind_nac   = report.get('industria_nacional_ytd', {})
    ford_nac  = report.get('ford_nacional_ytd', {})

    def dict_to_arr(d):
        """Convert {cat: {y2024,y2025,y2026}} to array with ytd2024/ytd2025/ytd2026 field names.
        page.tsx expects ytd2025/ytd2026. Includes _total as 'Total general' entry."""
        rows = []
        total = d.get('_total')
        for k, v in d.items():
            if k == '_total': continue
            rows.append({'cat': k, 'ytd2024': v.get('y2024'), 'ytd2025': v.get('y2025'),
                         'ytd2026': v.get('y2026'), 'fcts2026': v.get('fcts2026')})
        if total:
            rows.append({'cat': 'Total general', 'ytd2024': total.get('y2024'),
                         'ytd2025': total.get('y2025'), 'ytd2026': total.get('y2026'),
                         'fcts2026': total.get('fcts2026')})
        return rows

    out['mercado_ytd']          = dict_to_arr(ind_orgu)
    out['ford_ytd']             = dict_to_arr(ford_orgu)
    out['mercado_ytd_nacional'] = dict_to_arr(ind_nac)
    out['ford_ytd_nacional']    = dict_to_arr(ford_nac)

    # ── T1: cat por provincia individual ──
    prov_sheet_map = {
        'cat_pichincha_ytd': 'CAT_PICHINCHA_YTD',
        'cat_guayas_ytd':    'CAT_GUAYAS_YTD',
        'cat_manabi_ytd':    'CAT_MANABI_YTD',
        'cat_eloro_ytd':     'CAT_ELORO_YTD',
    }
    for key, sheet in prov_sheet_map.items():
        out[key] = dict_to_arr(extract_summary(wb[sheet]))

    # ── T1: provincias_ytd (array con label/ytd2024/ytd2025/ytd2026) ──
    out['provincias_ytd'] = [
        {'label': k, 'ytd2024': v.get('y2024'), 'ytd2025': v.get('y2025'), 'ytd2026': v.get('y2026')}
        for k, v in report.get('provincias_industria', {}).items()
    ]
    out['ford_provincias_ytd'] = [
        {'label': k, 'ytd2024': v.get('y2024'), 'ytd2025': v.get('y2025'), 'ytd2026': v.get('y2026')}
        for k, v in report.get('provincias_ford', {}).items()
    ]

    # ford_cat_por_prov: page.tsx espera array flat [{label:'PICHINCHA'}, {label:'SUV', ytd2025, ytd2026}, {label:'GUAYAS'}, ...]
    fcpp = report.get('ford_cat_por_prov', {})
    flat_fcpp = []
    for prov, cats in fcpp.items():
        flat_fcpp.append({'label': prov})
        for cat_name, vals in cats.items():
            flat_fcpp.append({
                'label': cat_name,
                'ytd2024': vals.get('y2024'), 'ytd2025': vals.get('y2025'), 'ytd2026': vals.get('y2026')
            })
    out['ford_cat_por_prov'] = flat_fcpp

    # ── T2: combustible SUV/PU ──
    def parse_combustible_nac(ws):
        rows = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))
        result = []
        for i, row in enumerate(rows):
            if row[0] == 'Etiquetas de fila':
                for r in rows[i+1:]:
                    if r[0] is None: break
                    if str(r[0]) == 'Total general': continue
                    if str(r[0]) in ('SUV', 'PICK UPS'): continue
                    result.append({'label': r[0], '2024': safe_num(r[1]),
                                   '2025': safe_num(r[2]), '2026': safe_num(r[3])})
        return result

    def combustible_prov_to_flat(ws):
        """Returns flat array: [{label:'PICHINCHA'}, {label:'GASOLINA','2024':...}, ...]
        page.tsx iterates this and uses r.label to detect province vs fuel."""
        rows = list(ws.iter_rows(min_row=1, max_row=80, values_only=True))
        result = []
        for i, row in enumerate(rows):
            if row[0] == 'Etiquetas de fila':
                for r in rows[i+1:]:
                    if r[0] is None: break
                    label = str(r[0]) if r[0] else ''
                    if not label or label == 'Total general': continue
                    if label in ('SUV','PICK UPS'): continue
                    result.append({'label': label, '2024': safe_num(r[1]),
                                   '2025': safe_num(r[2]), '2026': safe_num(r[3])})
        return result

    out['combustible_suv_nacional'] = parse_combustible_nac(wb['COMBUSTIBLE_NACIONAL_SUV_YTD'])
    out['combustible_suv_prov']     = combustible_prov_to_flat(wb['COMBUSTIBLE_PROV_SUV_YTD'])
    out['ford_comb_suv_nacional']   = parse_combustible_nac(wb['FORD_COMBUSTIBLE_SUV_YTD'])
    out['ford_comb_suv_prov']       = combustible_prov_to_flat(wb['FORD_COMBUSTIBLE_SUV_PROV_YTD'])
    out['combustible_pu_nacional']  = parse_combustible_nac(wb['COMBUSTIBLE_NACIONAL_PU_YTD'])
    out['combustible_pu_prov']      = combustible_prov_to_flat(wb['COMBUSTIBLE_PROV_PU_YTD'])
    out['ford_comb_pu_nacional']    = parse_combustible_nac(wb['FORD_COMBUSTIBLE_PU_YTD'])
    out['ford_comb_pu_prov']        = combustible_prov_to_flat(wb['FORD_COMBUSTIBLE_PU_PROV_YTD'])

    # ── T3: suv_segmentos, suv_segmentos_nacional, suv_seg_por_provincia ──
    def seg_dict_to_arr(d):
        return [{'seg': k, 'y2024': v.get('y2024'), 'y2025': v.get('y2025'),
                 'y2026': v.get('y2026'), 'fcts2026': v.get('fcts2026')} for k, v in d.items() if k != '_total']

    out['suv_segmentos']          = seg_dict_to_arr(report.get('suv_segmentos_orgu_ytd', {}))
    out['suv_segmentos_nacional'] = seg_dict_to_arr(extract_summary(wb['SUV_SEGMENTOS_NACIONAL_YTD']))
    out['ford_suv_seg_nacional']  = seg_dict_to_arr(extract_summary(wb['FORD_SUV_SEGMENTOS_NACIONAL_YTD']))

    # suv_seg_por_provincia: sheet has seg→province structure (B SUV→PICHINCHA,GUAYAS...)
    # page.tsx line 761 expects flat array: [{label:'PICHINCHA'},{label:'B SUV',y2024,...},...]
    # BUT actually page.tsx builds segGroups by iterating: if label in PROVS → new group key
    # So we need: [{label:'B SUV'}, {label:'PICHINCHA',y2024,...}, {label:'GUAYAS',...}, ...]
    def parse_seg_por_prov_flat(ws):
        rows = list(ws.iter_rows(min_row=1, max_row=100, values_only=True))
        result = []
        for i, row in enumerate(rows):
            if row[0] == 'Etiquetas de fila':
                year_cols = {str(h): j for j, h in enumerate(row) if str(h) in ('2024','2025','2026')}
                current_seg = None
                for r in rows[i+1:]:
                    if r[0] is None: break
                    label = str(r[0])
                    if label == 'Total general': continue
                    if label in PROVINCES:
                        # Province row under a segment
                        result.append({'label': label,
                            'y2024': safe_num(r[year_cols.get('2024', 99)]) if '2024' in year_cols else None,
                            'y2025': safe_num(r[year_cols.get('2025', 99)]) if '2025' in year_cols else None,
                            'y2026': safe_num(r[year_cols.get('2026', 99)]) if '2026' in year_cols else None,
                            '_seg': current_seg})
                    else:
                        # Segment header row
                        current_seg = label
                        result.append({'label': label})
                break
        return result

    out['suv_seg_por_provincia'] = parse_seg_por_prov_flat(wb['SUV_SEGMENTOS_POR_PROVINCIA_YTD'])

    ford_seg_prov_raw = extract_by_province(wb['FORD_SUV_SEGMENTOS_PROV_YTD'], brand_level=False)
    flat_ford_seg_prov = []
    for prov, rows in ford_seg_prov_raw.items():
        flat_ford_seg_prov.append({'label': prov})
        for r in rows:
            flat_ford_seg_prov.append({'label': r['label'], 'y2024': r.get('y2024'),
                                       'y2025': r.get('y2025'), 'y2026': r.get('y2026')})
    out['ford_suv_seg_prov'] = flat_ford_seg_prov

    # ── T4: suv_25_40_gas nested format ──
    # page.tsx: d.suv_25_40_gas.NACIONAL (array year rows), d.suv_25_40_gas.prov_marcas, d.suv_25_40_gas.por_provincia
    def build_national_year_rows(brand_list):
        """Convert [{brand, y2024, y2025, y2026}] → [{year:'2024', BRAND:val,...}, ...]"""
        result = {}
        for item in brand_list:
            for yr in ('y2024','y2025','y2026'):
                y = yr[1:]
                if y not in result: result[y] = {'year': y}
                result[y][item['brand']] = item.get(yr)
        return [result[y] for y in ('2024','2025','2026') if y in result]

    def build_prov_marcas(prov_dict):
        """Convert {PROV:[{brand,y2024,y2025,y2026}]} → {PROV:[{year,BRAND:val}]}"""
        out2 = {}
        for prov, brands in prov_dict.items():
            yr_rows = {}
            for item in brands:
                for yr in ('y2024','y2025','y2026'):
                    y = yr[1:]
                    if y not in yr_rows: yr_rows[y] = {'year': y}
                    yr_rows[y][item['brand']] = item.get(yr)
            out2[prov] = [yr_rows[y] for y in ('2024','2025','2026') if y in yr_rows]
        return out2

    def build_por_provincia(prov_vol_dict):
        """Convert {PROV:{y2024,y2025,y2026}} → [{label,y2024,y2025,y2026}]"""
        return [{'label': k, 'y2024': v.get('y2024'), 'y2025': v.get('y2025'), 'y2026': v.get('y2026')}
                for k, v in prov_vol_dict.items() if k != 'ZONA ORGU']

    out['suv_25_40_gas'] = {
        'NACIONAL':      build_national_year_rows(report.get('suv_gas_25_40_nacional', [])),
        'prov_marcas':   build_prov_marcas(report.get('suv_gas_25_40_por_prov', {})),
        'por_provincia': build_por_provincia(report.get('suv_gas_25_40_prov_vol', {})),
    }

    # ── T5: suv_25_40_fhev ──
    out['suv_25_40_fhev'] = {
        'NACIONAL':      build_national_year_rows(report.get('suv_hib_25_40_nacional', [])),
        'prov_marcas':   build_prov_marcas(report.get('suv_hib_25_40_por_prov', {})),
        'por_provincia': build_por_provincia(report.get('suv_hib_25_40_prov_vol', {})),
    }

    # ── T6: suv_40_50 ──
    out['suv_40_50'] = {
        'NACIONAL':      build_national_year_rows(report.get('suv_hib_40_50_nacional', [])),
        'prov_marcas':   build_prov_marcas(report.get('suv_hib_40_50_por_prov', {})),
        'por_provincia': build_por_provincia(report.get('suv_hib_40_50_prov_vol', {})),
    }

    # ── T7: suv_55_80 (Everest) and suv_60_80 (Explorer) ──
    out['suv_55_80'] = {
        'NACIONAL':      build_national_year_rows(report.get('suv_55_80_everest_nacional', [])),
        'prov_marcas':   build_prov_marcas(report.get('suv_55_80_por_prov', {})),
        'por_provincia': build_por_provincia({}),
    }
    out['suv_60_80'] = {
        'NACIONAL':      build_national_year_rows(report.get('suv_60_80_explorer_nacional', [])),
        'prov_marcas':   build_prov_marcas(report.get('suv_60_80_por_prov', {})),
        'por_provincia': build_por_provincia({}),
    }

    # ── T8: suv_80plus ──
    out['suv_80plus'] = {
        'NACIONAL':      build_national_year_rows(report.get('suv_80plus_nacional', [])),
        'prov_marcas':   build_prov_marcas(report.get('suv_80plus_por_prov', {})),
        'por_provincia': build_por_provincia({}),
    }

    # ── T9: pickup_cat_ytd, pickup_cat_nacional, pu_cat_por_prov ──
    out['pickup_cat_ytd']      = seg_dict_to_arr(extract_summary(wb['PICKUP_CAT_ORGU_YTD']))
    out['pickup_cat_nacional'] = seg_dict_to_arr(extract_summary(wb['PICKUP_CAT_NACIONAL_YTD']))

    def parse_pu_cat_por_prov(ws):
        # page.tsx line 2054 expects flat array with label as province or seg name
        rows = list(ws.iter_rows(min_row=1, max_row=50, values_only=True))
        result = []
        current_prov = None
        for row in rows:
            if row[0] in PROVINCES:
                current_prov = row[0]
                result.append({'label': current_prov})
                continue
            if current_prov and row[0] and row[0] not in ('Etiquetas de fila','Total general'):
                if row[0] in ('COMPACT PICK UPS','FULL SIZE PICK UPS','MID SIZE PICK UPS'):
                    result.append({'label': row[0], 'ytd2024': safe_num(row[1]),
                                   'ytd2025': safe_num(row[2]), 'ytd2026': safe_num(row[3])})
        return result

    out['pu_cat_por_prov'] = parse_pu_cat_por_prov(wb['PU_CAT_POR_PROV'])

    # pick_diesel.FORD: page.tsx T9 accesses d.pick_diesel.FORD as an array of year rows
    # Build from pu_diesel_ford flat data -> [{year:'2024', FORD:val}, ...]
    def build_ford_year_rows(brand_list, ford_key='FORD'):
        """Build [{year, FORD:val}] from [{brand, y2024, y2025, y2026}]"""
        ford_row = next((r for r in brand_list if r.get('brand') == ford_key), None)
        if not ford_row:
            return []
        return [
            {'year': '2024', ford_key: ford_row.get('y2024')},
            {'year': '2025', ford_key: ford_row.get('y2025')},
            {'year': '2026', ford_key: ford_row.get('y2026')},
        ]

    # ── T10: pick_diesel_tm / pick_diesel_ta ──
    prov_tm = extract_by_province(wb['PU_DIESEL_POR_PROV_MARCAS_TM'])
    prov_ta = extract_by_province(wb['PU_DIESEL_POR_PROV_MARCAS_TA'])
    out['pick_diesel_tm'] = {
        'NACIONAL':    build_national_year_rows(report.get('pu_diesel_tm_nacional', [])),
        'prov_marcas': build_prov_marcas(prov_tm),
        'FORD': build_ford_year_rows(report.get('pu_diesel_tm_nacional', [])),
    }
    out['pick_diesel_ta'] = {
        'NACIONAL':    build_national_year_rows(report.get('pu_diesel_ta_nacional', [])),
        'prov_marcas': build_prov_marcas(prov_ta),
        'FORD': build_ford_year_rows(report.get('pu_diesel_ta_nacional', [])),
    }

    # pick_diesel (combined, used in T9)
    combined_nac = []
    for brand_tm in report.get('pu_diesel_tm_nacional', []):
        brand_ta = next((b for b in report.get('pu_diesel_ta_nacional', []) if b['brand'] == brand_tm['brand']), {})
        combined_nac.append({
            'brand': brand_tm['brand'],
            'y2024': (brand_tm.get('y2024') or 0) + (brand_ta.get('y2024') or 0),
            'y2025': (brand_tm.get('y2025') or 0) + (brand_ta.get('y2025') or 0),
            'y2026': (brand_tm.get('y2026') or 0) + (brand_ta.get('y2026') or 0),
        })
    out['pick_diesel'] = {
        'NACIONAL':    build_national_year_rows(combined_nac),
        'prov_marcas': build_prov_marcas(report.get('pu_diesel_por_prov', {})),
        'FORD': build_ford_year_rows(combined_nac),
    }

    # ── T11: pick_fullsize ──
    out['pick_fullsize'] = {
        'NACIONAL':    build_national_year_rows(report.get('pu_fullsize_nacional', [])),
        'prov_marcas': build_prov_marcas(report.get('pu_fullsize_por_prov', {})),
        'FORD': build_ford_year_rows(report.get('pu_fullsize_nacional', [])),
    }

    # ── T12: ford_ytd already set above ──
    # mercado_ytd already set above

    return out

def main(excel_path):
    print(f"Leyendo: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    report = {}

    # ── Metadata ──
    meta_ws = wb['METADATA']
    meta_rows = list(meta_ws.iter_rows(max_row=20, values_only=True))
    meta = {str(r[0]): r[1] for r in meta_rows if r[0] and r[1]}
    months_ytd = int(meta.get('meses_incluidos', 5))
    mes_ytd = meta.get('mes_ytd')
    months_es = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                 'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    if hasattr(mes_ytd, 'month'):
        month_name = months_es[mes_ytd.month - 1] + ' ' + str(mes_ytd.year)
    else:
        month_name = 'Mayo 2026'
    report['report_month'] = month_name
    report['months_ytd'] = months_ytd
    print(f"Mes: {month_name} ({months_ytd} meses YTD)")

    # ── T1: Industria + Ford nacional FY ──
    print("  T1: Nacional FY/YTD...")
    ind_fy   = extract_summary(wb['IND_NACIONAL_FY'])
    ford_fy  = extract_summary(wb['FORD_NACIONAL_FY'])
    ind_ytd  = extract_summary(wb['IND_NACIONAL_YTD'])
    ford_ytd = extract_summary(wb['FORD_NACIONAL_YTD'])
    report['industria_nacional_fy']  = ind_fy
    report['ford_nacional_fy']       = ford_fy
    report['industria_nacional_ytd'] = ind_ytd
    report['ford_nacional_ytd']      = ford_ytd

    # ── T2: Zona Orgu FY/YTD ──
    print("  T2: Zona Orgu FY/YTD...")
    report['industria_orgu_fy']  = extract_summary(wb['IND_ORGU_FY'])
    report['ford_orgu_fy']       = extract_summary(wb['FORD_ORGU_FY'])
    report['industria_orgu_ytd'] = extract_summary(wb['IND_ORGU_YTD'])
    report['ford_orgu_ytd']      = extract_summary(wb['FORD_ORGU_YTD'])

    # ── T2: Provincias ──
    print("  T2: Provincias...")
    report['provincias_industria'] = extract_province_totals(wb['PROVINCIAS_YTD'])
    report['provincias_ford']      = extract_province_totals(wb['FORD_PROVINCIAS_YTD'])

    # ── T3: Categorías Orgu ──
    print("  T3: Categorías...")
    report['categorias_orgu_ytd']   = extract_summary(wb['CAT_ORGU_YTD'])
    report['categorias_orgu_fy']    = extract_summary(wb['CAT_ORGU_FY'])
    report['cat_por_provincia_ytd'] = extract_flat(wb['CAT_POR_PROVINCIA_YTD'])
    report['ford_cat_por_prov']     = extract_ford_cat_by_prov(wb['FORD_CATEGORIA_POR_PROV_YTD'])

    # ── T4: SUV Segmentos ──
    print("  T4: SUV Segmentos...")
    report['suv_segmentos_orgu_ytd'] = extract_summary(wb['SUV_SEGMENTOS_ORGU_YTD'])
    report['suv_segmentos_orgu_fy']  = extract_summary(wb['SUV_SEGMENTOS_ORGU_FY'])
    report['suv_segmentos_por_prov'] = extract_by_province(wb['SUV_SEGMENTOS_POR_PROVINCIA_YTD'], brand_level=False)

    # ── T5: SUV GAS 25-40K ──
    print("  T5: SUV GAS 25-40K...")
    report['suv_gas_25_40_nacional']  = extract_nacional_brands(wb['SUV_GAS_25_40_INDUSTRIA_YTD'])
    report['suv_gas_25_40_por_prov']  = extract_by_province(wb['SUV_GAS_25_40_POR_PROV_MARCAS'])
    report['suv_gas_25_40_ford']      = extract_flat(wb['SUV_GAS_25_40_FORD_YTD'])
    report['suv_gas_25_40_prov_vol']  = extract_province_totals(wb['SUV_GAS_25_40_POR_PROVINCIA_YTD'])

    # ── T6: SUV FHEV 25-40K ──
    print("  T6: SUV FHEV 25-40K...")
    report['suv_hib_25_40_nacional']  = extract_nacional_brands(wb['SUV_HIB_25_40_INDUSTRIA_YTD'])
    report['suv_hib_25_40_por_prov']  = extract_by_province(wb['SUV_HIB_25_40_POR_PROV_MARCAS'])
    report['suv_hib_25_40_ford']      = extract_flat(wb['SUV_HIB_25_40_FORD_YTD'])
    report['suv_hib_25_40_prov_vol']  = extract_province_totals(wb['SUV_HIB_25_40_POR_PROVINCIA_YTD'])

    # ── T7: SUV FHEV 40-50K ──
    print("  T7: SUV FHEV 40-50K...")
    report['suv_hib_40_50_nacional']  = extract_nacional_brands(wb['SUV_HIB_40_50_INDUSTRIA_YTD'])
    report['suv_hib_40_50_por_prov']  = extract_by_province(wb['SUV_HIB_40_50_POR_PROV_MARCAS'])
    report['suv_hib_40_50_ford']      = extract_flat(wb['SUV_HIB_40_50_FORD_YTD'])
    report['suv_hib_40_50_prov_vol']  = extract_province_totals(wb['SUV_HIB_40_50_POR_PROVINCIA'])

    # ── T8: SUV GAS 55-80K (Everest) ──
    print("  T8: SUV 55-80K Everest...")
    report['suv_55_80_everest_nacional'] = extract_nacional_brands(wb['SUV_GAS_55_80_EVEREST_INDUSTRIA'])
    report['suv_55_80_everest_ford']     = extract_flat(wb['SUV_GAS_55_80_EVEREST_FORD_YTD'])
    report['suv_55_80_por_prov']         = extract_by_province(wb['SUV_55_80_POR_PROV'])

    # ── T9: SUV GAS 60-80K (Explorer Active) ──
    print("  T9: SUV 60-80K Explorer...")
    report['suv_60_80_explorer_nacional'] = extract_nacional_brands(wb['SUV_GAS_55_80_EXPLORER_ACTIVE_I'])
    report['suv_60_80_explorer_ford']     = extract_flat(wb['SUV_GAS_55_80_EXPLORER_ACTIVE_F'])
    report['suv_60_80_por_prov']          = extract_by_province(wb['SUV_60_80_POR_PROV'])

    # ── T10: SUV 80K+ ──
    print("  T10: SUV 80K+...")
    report['suv_80plus_nacional']  = extract_nacional_brands(wb['SUV_80PLUS_INDUSTRIA_YTD'])
    report['suv_80plus_ford']      = extract_flat(wb['SUV_80PLUS_FORD_YTD'])
    report['suv_80plus_por_prov']  = extract_by_province(wb['SUV_80PLUS_POR_PROV'])

    # ── T11: Pick Ups 4x4 Diesel ──
    print("  T11: Pick Ups Diesel...")
    report['pu_diesel_tm_nacional']  = extract_nacional_brands(wb['PICKUP_4X4_DSL_50_70_IND_TM'])
    report['pu_diesel_ta_nacional']  = extract_nacional_brands(wb['PICKUP_4X4_DSL_50_70_IND_TA'])
    report['pu_diesel_por_prov']     = extract_pick_diesel_combined(
                                          wb['PU_DIESEL_POR_PROV_MARCAS_TM'],
                                          wb['PU_DIESEL_POR_PROV_MARCAS_TA'])
    report['pu_diesel_ford']         = extract_flat(wb['PICKUP_4X4_DSL_50_70_FORD_YTD'])
    report['pu_diesel_prov_vol_tm']  = extract_province_totals(wb['PICKUP_4X4_DSL_50_70_PROV_TM'])
    report['pu_diesel_prov_vol_ta']  = extract_province_totals(wb['PICKUP_4X4_DSL_50_70_PROV_TA'])

    # ── T12: Pick Ups Full Size ──
    print("  T12: Pick Ups Full Size...")
    report['pu_fullsize_nacional']  = extract_nacional_brands(wb['PICKUP_FULLSIZE_60_100_INDUSTRI'])
    report['pu_fullsize_por_prov']  = extract_by_province(wb['PU_FULLSIZE_POR_PROV_MARCAS'])
    report['pu_fullsize_ford']      = extract_flat(wb['PICKUP_FULLSIZE_60_100_FORD_YTD'])
    report['pu_fullsize_prov_vol']  = extract_province_totals(wb['PICKUP_FULLSIZE_60_100_POR_PROV'])

    # ── TOP 10 Modelos ──
    print("  Top 10 modelos...")
    report['top10_modelos'] = extract_flat(wb['TOP10_MODELOS_ORGU_YTD'])

    # ── Reshape for frontend compatibility ──
    print("  Reshaping para frontend...")
    report = reshape_for_frontend(report, wb)

    wb.close()

    # ── Write output ──
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'public', 'report_data.json')
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    # Preserve existing non-data keys (precios, bbc_hotlines, insights, etc.)
    existing = {}
    if os.path.exists(out_path):
        with open(out_path, 'r', encoding='utf-8') as f:
            existing = json.load(f)
        preserve_keys = ['precios_competidores', 'bbc_hotlines', 'model_filters',
                         'model_display_names', 'insights', 'vol_override', 'ford_cards']
        for k in preserve_keys:
            if k in existing:
                report[k] = existing[k]
        print(f"  Preservadas {len([k for k in preserve_keys if k in existing])} claves de admin.")

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)

    print(f"\n✅ Listo! → {out_path}")
    print(f"   Mes: {report['report_month']} ({months_ytd} meses YTD)")
    ind_total = ind_fy.get('_total', {})
    ford_total = ford_fy.get('_total', {})
    print(f"   Industria Nacional FY: 2024={ind_total.get('y2024')} | 2025={ind_total.get('y2025')} | 2026 YTD={ind_total.get('y2026')}")
    print(f"   Ford Nacional FY:      2024={ford_total.get('y2024')} | 2025={ford_total.get('y2025')} | 2026 YTD={ford_total.get('y2026')}")
    orgu_total = report['industria_orgu_fy'].get('_total', {})
    ford_orgu  = report['ford_orgu_fy'].get('_total', {})
    print(f"   Industria Orgu FY:     2024={orgu_total.get('y2024')} | 2025={orgu_total.get('y2025')} | 2026 YTD={orgu_total.get('y2026')}")
    print(f"   Ford Orgu FY:          2024={ford_orgu.get('y2024')} | 2025={ford_orgu.get('y2025')} | 2026 YTD={ford_orgu.get('y2026')}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python parser.py <ruta_excel.xlsx>")
        sys.exit(1)
    main(sys.argv[1])


# ─── RESHAPE: adapta el JSON al formato esperado por page.tsx ──────────────────
